"""
Excel workbook export.

Consolidates all non-empty CSV output files into a single .xlsx workbook,
with one sheet per CSV.  Sheets are ordered by investigative priority so
analysts land on the most actionable data first.
"""

import csv
from pathlib import Path

# Excel's hard limit is 1,048,576 rows per sheet (row 1 is the header).
_EXCEL_MAX_ROWS = 1_048_576

# Number of data rows sampled to size columns (avoids scanning huge sheets).
_WIDTH_SAMPLE_ROWS = 200

_INVALID_SHEET_CHARS = str.maketrans("[]:*?/\\", "_______")

# Sheets appear in this order; any CSV not in the list is appended alphabetically.
_SHEET_ORDER = [
    "alerts",
    "iocs",
    "timeline",
    "credential_findings",
    "credential_posts",
    "malicious_ja3",
    "malicious_ja4",
    "yara_hits",
    "arp_anomalies",
    "icmp_tunneling_candidates",
    "dns_tunneling_candidates",
    "beaconing_candidates",
    "entropy_exfil_candidates",
    "suspicious_downloads",
    "lateral_movement_candidates",
    "protocol_anomalies",
    "http_response_anomalies",
    "tls_sni_anomalies",
    "suspicious_user_agents",
    "file_indicators",
    "jarm_fingerprints",
    "ja4h",
    "os_fingerprints",
    "smtp_attachments",
    "smtp_activity",
    "dns_resolutions",
    "http_requests",
    "http_body_previews",
    "http_tshark",
    "http_responses",
    "tls_metadata",
    "carved_files",
    "extracted_payloads_index",
    "tcp_stream_index",
    "smb_tshark",
    "ftp_tshark",
    "kerberos_activity",
]

_HEADER_COLOR = "1F3864"   # dark navy
_ALT_ROW_COLOR = "EEF2F7"  # light blue-grey


def _sheet_name(stem: str) -> str:
    name = stem.translate(_INVALID_SHEET_CHARS)[:31]
    return name or "sheet"


def build_excel_workbook(output_dir: Path) -> Path | None:
    """
    Scan *output_dir* for CSV files, skip empty/header-only ones, and write
    all remaining data into an xlsx workbook.  Returns the workbook path, or
    None if openpyxl is not installed or no data was found.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    csv_files = {p.stem: p for p in output_dir.glob("*.csv")}
    if not csv_files:
        return None

    ordered = [s for s in _SHEET_ORDER if s in csv_files]
    remaining = sorted(s for s in csv_files if s not in set(ordered))
    ordered.extend(remaining)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=_HEADER_COLOR, end_color=_HEADER_COLOR,
                              fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    alt_fill = PatternFill(start_color=_ALT_ROW_COLOR, end_color=_ALT_ROW_COLOR,
                           fill_type="solid")

    sheets_written = 0
    for stem in ordered:
        csv_path = csv_files[stem]
        try:
            with open(csv_path, newline="", encoding="utf-8") as fh:
                rows = list(csv.reader(fh))
        except Exception:
            continue

        if not rows:
            continue
        data_rows = [r for r in rows[1:] if any(cell.strip() for cell in r)]
        if not data_rows:
            continue

        header = rows[0]

        # Excel can't hold more than _EXCEL_MAX_ROWS rows. If a CSV exceeds it
        # (e.g. a per-packet index on a large capture), keep as many as fit and
        # leave the final row as a pointer to the full CSV.
        omitted = 0
        if len(data_rows) > _EXCEL_MAX_ROWS - 1:
            keep = _EXCEL_MAX_ROWS - 2  # header row + one marker row
            omitted = len(data_rows) - keep
            data_rows = data_rows[:keep]

        ws = wb.create_sheet(title=_sheet_name(stem))

        for col_idx, value in enumerate(header, 1):
            cell = ws.cell(row=1, column=col_idx, value=value)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for row_idx, row in enumerate(data_rows, 2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if fill:
                    cell.fill = fill

        if omitted:
            ws.cell(
                row=len(data_rows) + 2,
                column=1,
                value=f"... {omitted} more rows omitted (Excel row limit) — see {stem}.csv",
            )

        # Size columns from the header plus a sample of data rows, rather than
        # scanning every cell (which is prohibitively slow on large sheets).
        widths = [len(str(h)) for h in header]
        for row in data_rows[:_WIDTH_SAMPLE_ROWS]:
            for i, value in enumerate(row):
                if i < len(widths):
                    widths[i] = max(widths[i], len(str(value)))
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 60)

        ws.freeze_panes = "A2"
        sheets_written += 1

    if sheets_written == 0:
        return None

    out_path = output_dir / "analysis_workbook.xlsx"
    wb.save(str(out_path))
    return out_path
